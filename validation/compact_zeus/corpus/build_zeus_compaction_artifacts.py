#!/usr/bin/env python3
"""Build shareable ZeusRawLogs compaction validation artifacts.

This is a validation-only utility. It reads the sample DataFrame with a
restricted pickle unpickler, applies compaction only below ``zeusRawLogs``,
and writes:

* a formatted Excel workbook with all sample rows and an omission audit;
* a standalone HTML report with six side-by-side comparisons;
* a Markdown technical introduction suitable for further editing.

No encoded content is decoded. The source DataFrame and ``agent_response``
values are never modified.
"""

from __future__ import annotations

import argparse
import html
import json
import re
import sys
from collections import Counter
from dataclasses import asdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from validation.compact_zeus.shared.compact_encoded_llm_context import (  # noqa: E402
    OmittedEncodedSpan,
    compact_zeus_raw_logs,
)
from validation.compact_zeus.shared.restricted_dataframe_pickle import (  # noqa: E402
    load_dataframe_pickle,
)

DEFAULT_DATA_PATH = (
    ROOT / "validation/compact_zeus/data/corpus/full_alert_validation_corpus.pkl"
)
DEFAULT_EXCEL_PATH = (
    ROOT
    / "validation/compact_zeus/data/compaction/zeus_raw_logs_compaction_report.xlsx"
)
DEFAULT_HTML_PATH = (
    ROOT
    / "validation/compact_zeus/data/compaction/zeus_raw_logs_compaction_technical_report.html"
)
DEFAULT_MARKDOWN_PATH = (
    ROOT / "validation/compact_zeus/docs/zeus_raw_logs_compaction_technical_intro.md"
)
KEY_ALERT_IDS = [1973909, 1981706, 1980288, 1979923, 1970506, 1979722]
MIN_BLOB_CHARS = 256

KIND_LABELS = {
    "pem": "PEM",
    "data_uri_base64": "Data URI Base64",
    "jwt_like": "JWT-like",
    "percent_encoded": "Percent-encoded",
    "hex_escape": "Hex escape",
    "unicode_escape": "Unicode escape",
    "hex_like": "Hex-like",
    "base64_like": "Base64-like",
}

KEY_ALERT_NOTES = {
    1973909: (
        "压缩量最大的重点样本。当前告警和 relatedAlertList 中存在重复的大段响应体、message、payload 及 JWT-like 内容，因此在不删除结构字段的前提下节省大量字符。"
    ),
    1981706: (
        "少量但极长的 Base64-like 片段占据绝大部分 alert_full_data，主要来自 message 和 http_http_response_body。"
    ),
    1980288: (
        "命中数量较多，但长编码只占完整告警的一部分；主要压缩 message、payload 中的 Base64-like/JWT-like 片段。"
    ),
    1979923: (
        "message 与 payload 中存在成对重复的长 Base64-like 片段，压缩后保留字段、路径、类型、原始长度和摘要。"
    ),
    1970506: (
        "命中字段较分散，包括 message、payload、请求/响应头和响应体；该样本也暴露了 http_http_method 被识别为 Base64-like 的语义误判候选，上线前需要以字段策略或评测集治理。"
    ),
    1979722: (
        "主要压缩 message、http_http_response_body 和 payload 中的 Base64-like 片段，其余告警结构保持不变。"
    ),
}


def load_sample(path: Path) -> pd.DataFrame:
    """Load the sample through the same restricted policy used by the notebooks."""

    return load_dataframe_pickle(path)


def compact_json_chars(value: Any) -> int:
    return len(
        json.dumps(
            value,
            ensure_ascii=False,
            separators=(",", ":"),
            default=str,
        )
    )


def scalar(value: Any) -> Any:
    """Convert numpy/pandas scalars to Excel- and JSON-friendly values."""

    if value is None:
        return ""
    if hasattr(value, "item") and callable(value.item):
        try:
            value = value.item()
        except (TypeError, ValueError):
            pass
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


def alert_id_value(value: Any) -> int | str:
    normalized = scalar(value)
    try:
        return int(normalized)
    except (TypeError, ValueError):
        return str(normalized)


def collect_zeus_nodes(
    value: Any,
    path: str = "$",
) -> list[tuple[str, Any]]:
    """Collect every zeusRawLogs container without descending into it twice."""

    result: list[tuple[str, Any]] = []
    if isinstance(value, dict):
        for key, child in value.items():
            child_path = child_path_for(path, str(key))
            if key == "zeusRawLogs":
                result.append((child_path, child))
            else:
                result.extend(collect_zeus_nodes(child, child_path))
    elif isinstance(value, list):
        for index, child in enumerate(value):
            result.extend(collect_zeus_nodes(child, f"{path}[{index}]"))
    return result


def child_path_for(path: str, key: str) -> str:
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", key):
        return f"{path}.{key}"
    return f"{path}[{json.dumps(key, ensure_ascii=False)}]"


def replace_zeus_nodes_with_sentinel(value: Any) -> Any:
    """Create a comparison projection proving non-Zeus values did not change."""

    if isinstance(value, dict):
        return {
            key: (
                "<ZEUS_RAW_LOGS_EXCLUDED_FROM_COMPARISON>"
                if key == "zeusRawLogs"
                else replace_zeus_nodes_with_sentinel(child)
            )
            for key, child in value.items()
        }
    if isinstance(value, list):
        return [replace_zeus_nodes_with_sentinel(child) for child in value]
    return value


def field_name_from_path(path: str) -> str:
    dot_match = re.search(r"\.([A-Za-z_][A-Za-z0-9_]*)$", path)
    if dot_match:
        return dot_match.group(1)
    bracket_match = re.search(r'\["((?:[^"\\]|\\.)*)"\]$', path)
    if bracket_match:
        try:
            return json.loads(f'"{bracket_match.group(1)}"')
        except json.JSONDecodeError:
            return bracket_match.group(1)
    return path.rsplit(".", maxsplit=1)[-1]


def source_area_from_path(path: str) -> str:
    if ".relatedAlertList[" in path:
        return "relatedAlertList"
    return "current_alert"


def marker_chars(omission: OmittedEncodedSpan) -> int:
    return len(
        f"<ENCODED:{omission.kind}:{omission.original_chars}:sha256={omission.sha256[:12]}:OMITTED>"
    )


def zeus_entry_count(nodes: list[tuple[str, Any]]) -> int:
    return sum(len(node) if isinstance(node, list) else 1 for _, node in nodes)


def process_sample(
    frame: pd.DataFrame,
    *,
    min_blob_chars: int,
) -> tuple[
    pd.DataFrame,
    pd.DataFrame,
    pd.DataFrame,
    dict[int, dict[str, Any]],
]:
    """Build row, field, audit, and key-comparison datasets."""

    row_records: list[dict[str, Any]] = []
    audit_records: list[dict[str, Any]] = []
    key_comparisons: dict[int, dict[str, Any]] = {}

    metadata_columns = [
        "topic",
        "topic_name",
        "risk_level",
        "status",
        "execute_type",
        "primary_type",
        "secondary_type",
        "tertiary_type",
        "predict_label",
        "ground_label",
    ]

    for _, row in frame.iterrows():
        alert_id = alert_id_value(row["alert_id"])
        source = row["alert_full_data"]
        if not isinstance(source, dict):
            raise TypeError(
                f"alert_id={alert_id}: alert_full_data must be dict, got {type(source)!r}"
            )

        compacted, omissions = compact_zeus_raw_logs(
            source,
            min_blob_chars=min_blob_chars,
        )
        if replace_zeus_nodes_with_sentinel(source) != replace_zeus_nodes_with_sentinel(
            compacted
        ):
            raise AssertionError(
                f"alert_id={alert_id}: a non-zeusRawLogs value changed"
            )
        if any(".zeusRawLogs" not in omission.path for omission in omissions):
            raise AssertionError(
                f"alert_id={alert_id}: omission escaped zeusRawLogs scope"
            )

        source_nodes = collect_zeus_nodes(source)
        compacted_nodes = dict(collect_zeus_nodes(compacted))
        if {path for path, _ in source_nodes} != set(compacted_nodes):
            raise AssertionError(
                f"alert_id={alert_id}: zeusRawLogs container paths changed"
            )

        before_chars = compact_json_chars(source)
        after_chars = compact_json_chars(compacted)
        saved_chars = before_chars - after_chars
        zeus_before_chars = sum(compact_json_chars(node) for _, node in source_nodes)
        zeus_after_chars = sum(
            compact_json_chars(compacted_nodes[path]) for path, _ in source_nodes
        )
        zeus_saved_chars = zeus_before_chars - zeus_after_chars
        omission_kinds = Counter(item.kind for item in omissions)
        unique_values = {item.sha256 for item in omissions}

        record: dict[str, Any] = {
            "alert_id": alert_id,
            **{
                column: scalar(row[column]) if column in frame.columns else ""
                for column in metadata_columns
            },
            "zeusRawLogs容器数": len(source_nodes),
            "zeusRawLogs日志条数": zeus_entry_count(source_nodes),
            "命中片段数": len(omissions),
            "唯一编码值数": len(unique_values),
            "Base64-like命中数": omission_kinds["base64_like"],
            "JWT-like命中数": omission_kinds["jwt_like"],
            "Percent-encoded命中数": omission_kinds["percent_encoded"],
            "其他类型命中数": len(omissions)
            - omission_kinds["base64_like"]
            - omission_kinds["jwt_like"]
            - omission_kinds["percent_encoded"],
            "alert_full_data压缩前字符数": before_chars,
            "alert_full_data压缩后字符数": after_chars,
            "alert_full_data节省字符数": saved_chars,
            "alert_full_data压缩率": (
                saved_chars / before_chars if before_chars else 0
            ),
            "zeusRawLogs压缩前字符数": zeus_before_chars,
            "zeusRawLogs压缩后字符数": zeus_after_chars,
            "zeusRawLogs节省字符数": zeus_saved_chars,
            "zeusRawLogs压缩率": (
                zeus_saved_chars / zeus_before_chars if zeus_before_chars else 0
            ),
            "非zeusRawLogs字段变化数": 0,
            "agent_response处理": "未处理",
        }
        row_records.append(record)

        for omission in omissions:
            audit_records.append(
                {
                    "alert_id": alert_id,
                    "topic": scalar(row.get("topic", "")),
                    "topic_name": scalar(row.get("topic_name", "")),
                    "来源区域": source_area_from_path(omission.path),
                    "字段": field_name_from_path(omission.path),
                    "编码类型": omission.kind,
                    "编码类型说明": KIND_LABELS.get(omission.kind, omission.kind),
                    "原始字符数": omission.original_chars,
                    "占位符字符数": marker_chars(omission),
                    "节省字符数": omission.original_chars - marker_chars(omission),
                    "sha256": omission.sha256,
                    "JSON路径": omission.path,
                }
            )

        if isinstance(alert_id, int) and alert_id in KEY_ALERT_IDS:
            changed_nodes = []
            for path, before_node in source_nodes:
                after_node = compacted_nodes[path]
                before_node_chars = compact_json_chars(before_node)
                after_node_chars = compact_json_chars(after_node)
                if before_node_chars == after_node_chars:
                    continue
                changed_nodes.append(
                    {
                        "path": path,
                        "before": before_node,
                        "after": after_node,
                        "before_chars": before_node_chars,
                        "after_chars": after_node_chars,
                        "saved_chars": before_node_chars - after_node_chars,
                    }
                )
            key_comparisons[alert_id] = {
                "row": record,
                "omissions": [asdict(item) for item in omissions],
                "changed_nodes": changed_nodes,
            }

    row_frame = pd.DataFrame(row_records)
    audit_frame = pd.DataFrame(audit_records)
    if audit_frame.empty:
        field_frame = pd.DataFrame(
            columns=[
                "alert_id",
                "topic",
                "topic_name",
                "来源区域",
                "字段",
                "编码类型",
                "编码类型说明",
                "命中片段数",
                "唯一编码值数",
                "原始字符数",
                "节省字符数",
            ]
        )
    else:
        field_frame = (
            audit_frame.groupby(
                [
                    "alert_id",
                    "topic",
                    "topic_name",
                    "来源区域",
                    "字段",
                    "编码类型",
                    "编码类型说明",
                ],
                dropna=False,
                as_index=False,
            )
            .agg(
                命中片段数=("sha256", "size"),
                唯一编码值数=("sha256", "nunique"),
                原始字符数=("原始字符数", "sum"),
                节省字符数=("节省字符数", "sum"),
            )
            .sort_values(
                ["alert_id", "原始字符数"],
                ascending=[True, False],
            )
        )
    return row_frame, field_frame, audit_frame, key_comparisons


def build_key_frame(
    row_frame: pd.DataFrame,
    field_frame: pd.DataFrame,
) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    row_lookup = row_frame.set_index("alert_id", drop=False)
    for alert_id in KEY_ALERT_IDS:
        if alert_id not in row_lookup.index:
            continue
        row = row_lookup.loc[alert_id]
        details = field_frame[field_frame["alert_id"] == alert_id].sort_values(
            "原始字符数",
            ascending=False,
        )
        compressed_parts = []
        for _, detail in details.head(8).iterrows():
            compressed_parts.append(
                f"{detail['字段']} / {detail['编码类型说明']}（{int(detail['命中片段数']):,} 段，{int(detail['原始字符数']):,} 字符）"
            )
        records.append(
            {
                "alert_id": alert_id,
                "topic": row["topic"],
                "命中片段数": int(row["命中片段数"]),
                "唯一编码值数": int(row["唯一编码值数"]),
                "压缩前字符数": int(row["alert_full_data压缩前字符数"]),
                "压缩后字符数": int(row["alert_full_data压缩后字符数"]),
                "节省字符数": int(row["alert_full_data节省字符数"]),
                "alert_full_data压缩率": float(row["alert_full_data压缩率"]),
                "zeusRawLogs压缩率": float(row["zeusRawLogs压缩率"]),
                "压缩了什么": "\n".join(compressed_parts),
                "样本说明": KEY_ALERT_NOTES.get(alert_id, ""),
            }
        )
    return pd.DataFrame(records)


def summarize(
    row_frame: pd.DataFrame,
    audit_frame: pd.DataFrame,
) -> dict[str, Any]:
    before = int(row_frame["alert_full_data压缩前字符数"].sum())
    after = int(row_frame["alert_full_data压缩后字符数"].sum())
    saved = before - after
    zeus_before = int(row_frame["zeusRawLogs压缩前字符数"].sum())
    zeus_after = int(row_frame["zeusRawLogs压缩后字符数"].sum())
    zeus_saved = zeus_before - zeus_after
    return {
        "rows": len(row_frame),
        "hit_rows": int((row_frame["命中片段数"] > 0).sum()),
        "matches": int(row_frame["命中片段数"].sum()),
        "unique_values": int(audit_frame["sha256"].nunique())
        if not audit_frame.empty
        else 0,
        "before_chars": before,
        "after_chars": after,
        "saved_chars": saved,
        "reduction_rate": saved / before if before else 0,
        "zeus_before_chars": zeus_before,
        "zeus_after_chars": zeus_after,
        "zeus_saved_chars": zeus_saved,
        "zeus_reduction_rate": (zeus_saved / zeus_before if zeus_before else 0),
        "non_zeus_changes": int(row_frame["非zeusRawLogs字段变化数"].sum()),
    }


def add_table(
    worksheet: Any,
    *,
    name: str,
    min_row: int,
    max_row: int,
    max_col: int,
) -> None:
    if max_row <= min_row:
        return
    reference = f"A{min_row}:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def style_sheet(
    worksheet: Any,
    *,
    header_row: int = 1,
    freeze_panes: str | None = "A2",
    wrap_columns: set[str] | None = None,
) -> None:
    wrap_columns = wrap_columns or set()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    for cell in worksheet[header_row]:
        if cell.value is None:
            continue
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=thin_gray)
    worksheet.freeze_panes = freeze_panes
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column_index)
        header = worksheet.cell(header_row, column_index).value
        max_length = len(str(header or ""))
        for row_index in range(
            header_row + 1,
            min(worksheet.max_row, header_row + 300) + 1,
        ):
            value = worksheet.cell(row_index, column_index).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        width = min(max(max_length + 2, 11), 55)
        if str(header) in wrap_columns:
            width = 55
        worksheet.column_dimensions[letter].width = width
        if str(header) in wrap_columns:
            for row_index in range(header_row + 1, worksheet.max_row + 1):
                worksheet.cell(row_index, column_index).alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )


def apply_number_formats(
    worksheet: Any,
    *,
    header_row: int = 1,
) -> None:
    percent_headers = {
        "alert_full_data压缩率",
        "zeusRawLogs压缩率",
    }
    integer_suffixes = ("字符数", "片段数", "编码值数", "变化数")
    for column_index in range(1, worksheet.max_column + 1):
        header = str(worksheet.cell(header_row, column_index).value or "")
        if header in percent_headers:
            number_format = "0.00%"
        elif header.endswith(integer_suffixes):
            number_format = "#,##0"
        else:
            continue
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            worksheet.cell(row_index, column_index).number_format = number_format


def column_index_by_header(
    worksheet: Any,
    header: str,
    *,
    header_row: int = 1,
) -> int:
    for column_index in range(1, worksheet.max_column + 1):
        if worksheet.cell(header_row, column_index).value == header:
            return column_index
    raise KeyError(f"header not found: {header}")


def write_excel(
    output_path: Path,
    *,
    source_path: Path,
    row_frame: pd.DataFrame,
    field_frame: pd.DataFrame,
    audit_frame: pd.DataFrame,
    key_frame: pd.DataFrame,
    summary: dict[str, Any],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    overview = workbook.active
    overview.title = "概览"

    overview.merge_cells("A1:F1")
    overview["A1"] = "ZeusRawLogs 长编码压缩技术验证"
    overview["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    overview["A1"].fill = PatternFill("solid", fgColor="17365D")
    overview["A1"].alignment = Alignment(horizontal="center")
    overview.row_dimensions[1].height = 30

    summary_rows = [
        ("指标", "结果", "说明"),
        ("源数据", str(source_path.relative_to(ROOT)), "受限反序列化读取"),
        (
            "实际样本行数",
            summary["rows"],
            "按实际输入 DataFrame 统计；文件名中的数量不作为校验口径",
        ),
        ("命中告警数", summary["hit_rows"], "至少命中 1 个长编码片段"),
        ("命中片段数", summary["matches"], "可包含同一编码值的重复出现"),
        ("唯一编码值数", summary["unique_values"], "按编码原文 SHA-256 去重"),
        (
            "alert_full_data 压缩前字符数",
            summary["before_chars"],
            "紧凑 JSON 字符数，不等同于模型 tokenizer token 数",
        ),
        (
            "alert_full_data 压缩后字符数",
            summary["after_chars"],
            "仅 zeusRawLogs 内编码片段被占位符替换",
        ),
        ("节省字符数", summary["saved_chars"], ""),
        (
            "alert_full_data 总体压缩率",
            summary["reduction_rate"],
            "节省字符数 / 压缩前字符数",
        ),
        (
            "zeusRawLogs 自身压缩率",
            summary["zeus_reduction_rate"],
            "只计算所有 zeusRawLogs 容器",
        ),
        ("非 zeusRawLogs 字段变化数", summary["non_zeus_changes"], "断言为 0"),
        ("agent_response", "未处理", "本验证不读取、不压缩、不改写"),
        ("是否解码", "否", "只识别编码形态，不尝试 Base64/JWT 等解码"),
        (
            "占位符",
            "<ENCODED:type:length:sha256=short-hash:OMITTED>",
            "保留类型、原始长度和短哈希；审计表另存路径和完整 SHA-256",
        ),
    ]
    for row_index, values in enumerate(summary_rows, start=3):
        for column_index, value in enumerate(values, start=1):
            overview.cell(row_index, column_index, value=value)
    overview["E3"] = "验证边界"
    overview["E4"] = "处理范围"
    overview["F4"] = "alert_full_data 下任意层级的 zeusRawLogs"
    overview["E5"] = "最小通用片段"
    overview["F5"] = f"{MIN_BLOB_CHARS} 字符"
    overview["E6"] = "原始数据"
    overview["F6"] = "不修改；Excel/HTML 为验证产物"
    overview["E7"] = "敏感信息"
    overview["F7"] = "HTML 左右对比保留原文，仅限授权内部分享"
    overview["E8"] = "生产结论"
    overview["F8"] = "当前仅为离线验证，不代表可直接接入 Runtime"

    for cell in overview[3]:
        if cell.column <= 3:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
    for row_index in range(3, 18):
        overview.cell(row_index, 1).font = Font(bold=row_index != 3)
        overview.cell(row_index, 1).alignment = Alignment(wrap_text=True)
        overview.cell(row_index, 2).alignment = Alignment(wrap_text=True)
        overview.cell(row_index, 3).alignment = Alignment(wrap_text=True)
    for row_index in range(3, 9):
        overview.cell(row_index, 5).fill = PatternFill("solid", fgColor="D9EAF7")
        overview.cell(row_index, 5).font = Font(bold=True)
        overview.cell(row_index, 6).alignment = Alignment(wrap_text=True)
    overview["B12"].number_format = "0.00%"
    overview["B13"].number_format = "0.00%"
    for row_index in range(4, 18):
        if isinstance(overview.cell(row_index, 2).value, int):
            overview.cell(row_index, 2).number_format = "#,##0"
    overview.column_dimensions["A"].width = 31
    overview.column_dimensions["B"].width = 29
    overview.column_dimensions["C"].width = 48
    overview.column_dimensions["D"].width = 3
    overview.column_dimensions["E"].width = 20
    overview.column_dimensions["F"].width = 52
    overview.freeze_panes = "A3"

    full_sheet = workbook.create_sheet("完整压缩率")
    append_frame(full_sheet, row_frame)
    style_sheet(full_sheet)
    apply_number_formats(full_sheet)
    add_table(
        full_sheet,
        name="FullCompressionRate",
        min_row=1,
        max_row=full_sheet.max_row,
        max_col=full_sheet.max_column,
    )
    rate_col = column_index_by_header(full_sheet, "alert_full_data压缩率")
    saved_col = column_index_by_header(full_sheet, "alert_full_data节省字符数")
    full_sheet.conditional_formatting.add(
        f"{get_column_letter(rate_col)}2:{get_column_letter(rate_col)}{full_sheet.max_row}",
        ColorScaleRule(
            start_type="min",
            start_color="FFFFFF",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF2CC",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    full_sheet.conditional_formatting.add(
        f"{get_column_letter(saved_col)}2:{get_column_letter(saved_col)}{full_sheet.max_row}",
        DataBarRule(
            start_type="num",
            start_value=0,
            end_type="max",
            color="5B9BD5",
        ),
    )

    field_sheet = workbook.create_sheet("字段明细")
    append_frame(field_sheet, field_frame)
    style_sheet(field_sheet)
    apply_number_formats(field_sheet)
    add_table(
        field_sheet,
        name="FieldDetails",
        min_row=1,
        max_row=field_sheet.max_row,
        max_col=field_sheet.max_column,
    )

    audit_sheet = workbook.create_sheet("命中审计")
    append_frame(audit_sheet, audit_frame)
    style_sheet(
        audit_sheet,
        wrap_columns={"JSON路径", "sha256"},
    )
    apply_number_formats(audit_sheet)
    add_table(
        audit_sheet,
        name="OmissionAudit",
        min_row=1,
        max_row=audit_sheet.max_row,
        max_col=audit_sheet.max_column,
    )

    key_sheet = workbook.create_sheet("重点对比")
    append_frame(key_sheet, key_frame)
    style_sheet(
        key_sheet,
        wrap_columns={"压缩了什么", "样本说明"},
    )
    apply_number_formats(key_sheet)
    add_table(
        key_sheet,
        name="KeyComparisons",
        min_row=1,
        max_row=key_sheet.max_row,
        max_col=key_sheet.max_column,
    )
    key_sheet.row_dimensions[1].height = 34
    for row_index in range(2, key_sheet.max_row + 1):
        key_sheet.row_dimensions[row_index].height = 112

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "重点告警节省字符数"
    chart.y_axis.title = "Alert ID"
    chart.x_axis.title = "节省字符数"
    saved_column = column_index_by_header(key_sheet, "节省字符数")
    alert_column = column_index_by_header(key_sheet, "alert_id")
    data = Reference(
        key_sheet,
        min_col=saved_column,
        min_row=1,
        max_row=key_sheet.max_row,
    )
    categories = Reference(
        key_sheet,
        min_col=alert_column,
        min_row=2,
        max_row=key_sheet.max_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 9
    chart.width = 16
    key_sheet.add_chart(chart, "M2")

    temp_path = output_path.with_name(f".{output_path.stem}.tmp{output_path.suffix}")
    workbook.save(temp_path)
    temp_path.replace(output_path)


def append_frame(worksheet: Any, frame: pd.DataFrame) -> None:
    worksheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        worksheet.append(
            [
                value.item()
                if hasattr(value, "item") and callable(value.item)
                else value
                for value in row
            ]
        )


def format_int(value: int | float) -> str:
    return f"{int(value):,}"


def format_percent(value: float) -> str:
    return f"{value * 100:.2f}%"


def render_json(value: Any, *, highlight_markers: bool = False) -> str:
    rendered = html.escape(json.dumps(value, ensure_ascii=False, indent=2, default=str))
    if highlight_markers:
        rendered = re.sub(
            r"(&lt;ENCODED:[^&\n]+?:\d+:OMITTED&gt;)",
            r"<mark>\1</mark>",
            rendered,
        )
    return rendered


def html_table(
    frame: pd.DataFrame,
    *,
    percent_columns: set[str] | None = None,
    integer_columns: set[str] | None = None,
    table_id: str | None = None,
) -> str:
    percent_columns = percent_columns or set()
    integer_columns = integer_columns or set()
    identifier = f' id="{html.escape(table_id)}"' if table_id else ""
    parts = [f"<table{identifier}><thead><tr>"]
    parts.extend(f"<th>{html.escape(str(column))}</th>" for column in frame.columns)
    parts.append("</tr></thead><tbody>")
    for _, row in frame.iterrows():
        parts.append("<tr>")
        for column in frame.columns:
            value = row[column]
            if column == "alert_id":
                try:
                    numeric_id = float(value)
                    text = (
                        str(int(numeric_id)) if numeric_id.is_integer() else str(value)
                    )
                except (TypeError, ValueError):
                    text = str(value)
            elif column in percent_columns:
                text = format_percent(float(value))
            elif column in integer_columns:
                text = format_int(value)
            else:
                text = str(value)
            parts.append(f"<td>{html.escape(text)}</td>")
        parts.append("</tr>")
    parts.append("</tbody></table>")
    return "".join(parts)


def top_field_frame(audit_frame: pd.DataFrame) -> pd.DataFrame:
    if audit_frame.empty:
        return pd.DataFrame()
    return (
        audit_frame.groupby(
            ["字段", "编码类型说明"],
            as_index=False,
        )
        .agg(
            命中片段数=("sha256", "size"),
            唯一编码值数=("sha256", "nunique"),
            原始字符数=("原始字符数", "sum"),
            节省字符数=("节省字符数", "sum"),
        )
        .sort_values("原始字符数", ascending=False)
        .head(20)
    )


def kind_summary_frame(audit_frame: pd.DataFrame) -> pd.DataFrame:
    if audit_frame.empty:
        return pd.DataFrame()
    return (
        audit_frame.groupby(
            ["编码类型", "编码类型说明"],
            as_index=False,
        )
        .agg(
            命中片段数=("sha256", "size"),
            唯一编码值数=("sha256", "nunique"),
            原始字符数=("原始字符数", "sum"),
            节省字符数=("节省字符数", "sum"),
        )
        .sort_values("原始字符数", ascending=False)
    )


def build_comparison_html(
    *,
    alert_id: int,
    comparison: dict[str, Any],
    field_frame: pd.DataFrame,
) -> str:
    row = comparison["row"]
    details = field_frame[field_frame["alert_id"] == alert_id][
        [
            "来源区域",
            "字段",
            "编码类型说明",
            "命中片段数",
            "唯一编码值数",
            "原始字符数",
            "节省字符数",
        ]
    ]
    detail_table = html_table(
        details,
        integer_columns={
            "命中片段数",
            "唯一编码值数",
            "原始字符数",
            "节省字符数",
        },
    )
    node_parts: list[str] = []
    for index, node in enumerate(comparison["changed_nodes"], start=1):
        node_parts.append(
            f"""
            <details class="node" {"open" if alert_id == 1973909 and index == 1 else ""}>
              <summary>
                zeusRawLogs #{index} · {html.escape(node["path"])}
                · {format_int(node["before_chars"])} → {format_int(node["after_chars"])}
                · 节省 {format_int(node["saved_chars"])}
              </summary>
              <div class="compare-grid">
                <section class="panel raw">
                  <h4>压缩前（原始 zeusRawLogs）</h4>
                  <pre>{render_json(node["before"])}</pre>
                </section>
                <section class="panel compacted">
                  <h4>压缩后（仅长编码片段替换）</h4>
                  <pre>{render_json(node["after"], highlight_markers=True)}</pre>
                </section>
              </div>
            </details>
            """
        )
    note = KEY_ALERT_NOTES.get(alert_id, "")
    return f"""
      <details class="alert-comparison" {"open" if alert_id == 1973909 else ""}>
        <summary>
          <strong>Alert {alert_id}</strong>
          <span>{format_int(row["alert_full_data压缩前字符数"])}
          → {format_int(row["alert_full_data压缩后字符数"])}</span>
          <span class="saved">节省 {format_int(row["alert_full_data节省字符数"])}
         （{format_percent(row["alert_full_data压缩率"])}）</span>
        </summary>
        <p class="note">{html.escape(note)}</p>
        <h3>压缩了什么</h3>
        <div class="table-scroll">{detail_table}</div>
        <h3>zeusRawLogs 左右对比</h3>
        <p class="muted">黄色标记为压缩后的占位符；点击每个容器展开。</p>
        {"".join(node_parts)}
      </details>
    """


def write_html(
    output_path: Path,
    *,
    source_path: Path,
    row_frame: pd.DataFrame,
    field_frame: pd.DataFrame,
    audit_frame: pd.DataFrame,
    key_frame: pd.DataFrame,
    key_comparisons: dict[int, dict[str, Any]],
    summary: dict[str, Any],
) -> None:
    key_display = key_frame[
        [
            "alert_id",
            "命中片段数",
            "唯一编码值数",
            "压缩前字符数",
            "压缩后字符数",
            "节省字符数",
            "alert_full_data压缩率",
            "zeusRawLogs压缩率",
        ]
    ]
    full_display = row_frame[
        [
            "alert_id",
            "topic",
            "risk_level",
            "命中片段数",
            "唯一编码值数",
            "alert_full_data压缩前字符数",
            "alert_full_data压缩后字符数",
            "alert_full_data节省字符数",
            "alert_full_data压缩率",
            "zeusRawLogs压缩率",
        ]
    ]
    kind_frame = kind_summary_frame(audit_frame)
    fields = top_field_frame(audit_frame)
    comparison_sections = "".join(
        build_comparison_html(
            alert_id=alert_id,
            comparison=key_comparisons[alert_id],
            field_frame=field_frame,
        )
        for alert_id in KEY_ALERT_IDS
        if alert_id in key_comparisons
    )
    document = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>ZeusRawLogs 长编码压缩技术验证</title>
  <style>
    :root {{
      --ink: #172033;
      --muted: #667085;
      --line: #d9e2ec;
      --blue: #175cd3;
      --blue-soft: #eff8ff;
      --green: #027a48;
      --green-soft: #ecfdf3;
      --orange: #b54708;
      --orange-soft: #fffaeb;
      --red: #b42318;
      --red-soft: #fef3f2;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      color: var(--ink);
      background: #f6f8fb;
      font: 15px/1.6 -apple-system, BlinkMacSystemFont, "Segoe UI",
            "PingFang SC", "Microsoft YaHei", sans-serif;
    }}
    main {{ max-width: 1540px; margin: 0 auto; padding: 36px 28px 80px; }}
    h1 {{ margin: 0 0 8px; font-size: 32px; }}
    h2 {{ margin: 42px 0 14px; font-size: 23px; }}
    h3 {{ margin: 24px 0 10px; }}
    h4 {{ margin: 0; padding: 11px 14px; font-size: 14px; }}
    p {{ margin: 8px 0; }}
    code {{
      padding: 1px 5px;
      border-radius: 4px;
      background: #eef2f6;
      font-family: ui-monospace, SFMono-Regular, Menlo, monospace;
    }}
    .subtitle {{ color: var(--muted); margin-bottom: 22px; }}
    .warning {{
      border-left: 5px solid var(--red);
      padding: 12px 16px;
      background: var(--red-soft);
      color: #7a271a;
      border-radius: 8px;
      margin: 20px 0;
    }}
    .cards {{
      display: grid;
      grid-template-columns: repeat(5, minmax(160px, 1fr));
      gap: 12px;
    }}
    .card {{
      padding: 16px;
      border: 1px solid var(--line);
      border-radius: 12px;
      background: white;
      box-shadow: 0 2px 8px rgba(16, 24, 40, .04);
    }}
    .card .label {{ color: var(--muted); font-size: 13px; }}
    .card .value {{ margin-top: 4px; font-size: 25px; font-weight: 700; }}
    .scope {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 14px;
    }}
    .scope section {{
      background: white;
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 16px 18px;
    }}
    .scope h3 {{ margin-top: 0; color: var(--blue); }}
    table {{
      width: 100%;
      border-collapse: separate;
      border-spacing: 0;
      background: white;
      font-size: 13px;
    }}
    th, td {{
      padding: 9px 10px;
      border-right: 1px solid var(--line);
      border-bottom: 1px solid var(--line);
      text-align: left;
      vertical-align: top;
      white-space: nowrap;
    }}
    th {{
      position: sticky;
      top: 0;
      z-index: 1;
      color: white;
      background: #1f4e78;
      cursor: pointer;
    }}
    tr:nth-child(even) td {{ background: #f8fafc; }}
    th:first-child, td:first-child {{ border-left: 1px solid var(--line); }}
    thead tr:first-child th {{ border-top: 1px solid var(--line); }}
    th:first-child {{ border-top-left-radius: 8px; }}
    th:last-child {{ border-top-right-radius: 8px; }}
    .table-scroll {{
      max-height: 560px;
      overflow: auto;
      border-radius: 8px;
      box-shadow: 0 0 0 1px var(--line);
    }}
    .alert-comparison {{
      margin: 14px 0;
      border: 1px solid var(--line);
      border-radius: 12px;
      background: white;
      overflow: hidden;
    }}
    .alert-comparison > summary {{
      display: flex;
      flex-wrap: wrap;
      gap: 18px;
      align-items: center;
      padding: 15px 18px;
      cursor: pointer;
      background: var(--blue-soft);
      font-size: 16px;
    }}
    .alert-comparison > summary .saved {{
      color: var(--green);
      font-weight: 700;
    }}
    .alert-comparison > p,
    .alert-comparison > h3,
    .alert-comparison > .table-scroll,
    .alert-comparison > .node {{ margin-left: 18px; margin-right: 18px; }}
    .note {{
      border-left: 4px solid var(--blue);
      padding: 10px 12px;
      background: #f8fafc;
    }}
    .node {{
      margin-top: 12px;
      margin-bottom: 16px;
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: hidden;
    }}
    .node > summary {{
      padding: 10px 12px;
      cursor: pointer;
      color: var(--blue);
      background: #f8fafc;
      font-family: ui-monospace, SFMono-Regular, Menlo, monospace;
      overflow-wrap: anywhere;
    }}
    .compare-grid {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
      gap: 1px;
      background: var(--line);
    }}
    .panel {{ min-width: 0; background: white; }}
    .panel.raw h4 {{ color: var(--orange); background: var(--orange-soft); }}
    .panel.compacted h4 {{ color: var(--green); background: var(--green-soft); }}
    pre {{
      max-height: 620px;
      margin: 0;
      padding: 14px;
      overflow: auto;
      white-space: pre-wrap;
      overflow-wrap: anywhere;
      background: #101828;
      color: #e4e7ec;
      font: 12px/1.5 ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
    }}
    mark {{
      padding: 1px 2px;
      color: #101828;
      background: #fef08a;
      border-radius: 3px;
    }}
    .muted {{ color: var(--muted); }}
    .search {{
      width: min(520px, 100%);
      margin: 0 0 12px;
      padding: 10px 12px;
      border: 1px solid #98a2b3;
      border-radius: 8px;
      font: inherit;
    }}
    footer {{ margin-top: 50px; color: var(--muted); font-size: 13px; }}
    @media (max-width: 1000px) {{
      .cards {{ grid-template-columns: repeat(2, 1fr); }}
      .scope, .compare-grid {{ grid-template-columns: 1fr; }}
    }}
    @media print {{
      body {{ background: white; }}
      main {{ max-width: none; padding: 10mm; }}
      .node:not([open]) {{ display: none; }}
      pre {{ max-height: none; }}
    }}
  </style>
</head>
<body>
<main>
  <h1>ZeusRawLogs 长编码压缩技术验证</h1>
  <p class="subtitle">
    目标：在不解码、不改原始告警的前提下，减少送入 LLM 的无效长编码字符。
    数据源：<code>{html.escape(str(source_path.relative_to(ROOT)))}</code>
  </p>
  <div class="warning">
    <strong>内部敏感材料：</strong>
    左右对比按需求保留原始 zeusRawLogs，可能包含凭证、Token、Cookie 或其他敏感内容。
    仅向有告警访问权限的同事分享。当前结果是离线验证，不代表已经接入 Runtime。
  </div>

  <div class="cards">
    <div class="card"><div class="label">实际样本</div>
      <div class="value">{format_int(summary["rows"])} 条</div></div>
    <div class="card"><div class="label">命中告警</div>
      <div class="value">{format_int(summary["hit_rows"])} 条</div></div>
    <div class="card"><div class="label">命中片段</div>
      <div class="value">{format_int(summary["matches"])}</div></div>
    <div class="card"><div class="label">节省字符</div>
      <div class="value">{format_int(summary["saved_chars"])}</div></div>
    <div class="card"><div class="label">alert_full_data 总体压缩率</div>
      <div class="value">{format_percent(summary["reduction_rate"])}</div></div>
  </div>
  <p class="muted">
    本报告按实际输入 DataFrame 统计并完整保留全部 {summary["rows"]} 行；
    文件名中的数量不作为校验口径。字符压缩率是紧凑 JSON 字符数口径，
    不是某个特定模型 tokenizer 的精确 token 压缩率。
  </p>

  <h2>技术方案与边界</h2>
  <div class="scope">
    <section>
      <h3>做什么</h3>
      <ul>
        <li>递归寻找 <code>alert_full_data</code> 下任意层级的
            <code>zeusRawLogs</code>。</li>
        <li>识别长度达到 {MIN_BLOB_CHARS} 字符的 Base64-like、Hex-like，
            以及 JWT-like、Percent-encoded、PEM 等连续编码形态。</li>
        <li>用 <code>&lt;ENCODED:type:length:sha256=short-hash:OMITTED&gt;</code> 替换命中片段。</li>
        <li>在审计明细中保留 JSON 路径、类型、原始长度和 SHA-256。</li>
      </ul>
    </section>
    <section>
      <h3>不做什么</h3>
      <ul>
        <li>不解码编码内容，不依据解码后的文件后缀或语义判断。</li>
        <li>不处理 <code>agent_response</code>，也不处理其他字段。</li>
        <li>不修改源 pickle 或原始 <code>alert_full_data</code>。</li>
        <li>不把这次离线字符统计等同于生产 token 成本或安全上线结论。</li>
      </ul>
    </section>
  </div>

  <h2>总体命中构成</h2>
  <div class="scope">
    <section>
      <h3>按编码类型</h3>
      <div class="table-scroll">
        {
        html_table(
            kind_frame,
            integer_columns={"命中片段数", "唯一编码值数", "原始字符数", "节省字符数"},
        )
    }
      </div>
    </section>
    <section>
      <h3>字符量最大的字段（Top 20）</h3>
      <div class="table-scroll">
        {
        html_table(
            fields,
            integer_columns={"命中片段数", "唯一编码值数", "原始字符数", "节省字符数"},
        )
    }
      </div>
    </section>
  </div>

  <h2>重点 6 条效果</h2>
  <div class="table-scroll">
    {
        html_table(
            key_display,
            percent_columns={"alert_full_data压缩率", "zeusRawLogs压缩率"},
            integer_columns={
                "命中片段数",
                "唯一编码值数",
                "压缩前字符数",
                "压缩后字符数",
                "节省字符数",
            },
        )
    }
  </div>

  <h2>重点告警：压缩内容与左右对比</h2>
  <p class="muted">
    默认展开 1973909，其余告警点击标题展开。每个 zeusRawLogs 容器也可单独展开，
    便于会议演示时控制页面长度。
  </p>
  {comparison_sections}

  <h2>全部 {summary["rows"]} 条压缩率</h2>
  <input class="search" id="full-search"
         placeholder="筛选 Alert ID、topic、risk_level……">
  <div class="table-scroll">
    {
        html_table(
            full_display,
            percent_columns={"alert_full_data压缩率", "zeusRawLogs压缩率"},
            integer_columns={
                "命中片段数",
                "唯一编码值数",
                "alert_full_data压缩前字符数",
                "alert_full_data压缩后字符数",
                "alert_full_data节省字符数",
            },
            table_id="full-table",
        )
    }
  </div>

  <h2>风险和下一步验证</h2>
  <div class="scope">
    <section>
      <h3>主要风险</h3>
      <ul>
        <li><strong>误压缩：</strong>长普通文本也可能碰巧满足编码字符集；
            1970506 的 <code>http_http_method</code> 命中值得单独复核。</li>
        <li><strong>漏压缩：</strong>短片段、分段编码、混杂特殊字符的编码可能未命中。</li>
        <li><strong>调查信息损失：</strong>LLM 看不到被替换原文，只能看到类型、长度；
            原文仍必须留在受控审计/运营界面。</li>
        <li><strong>凭证暴露：</strong>本 HTML 为展示效果保留原文，不应外发。</li>
      </ul>
    </section>
    <section>
      <h3>上线前建议度量</h3>
      <ul>
        <li>人工抽检命中片段的误压缩率和漏压缩率，按字段/Topic 分层。</li>
        <li>用目标模型 tokenizer 复测 token 节省率，而非只看字符数。</li>
        <li>比较压缩前后告警研判结论、证据引用和人工复核耗时。</li>
        <li>建立字段例外/允许策略和可回溯审计，再评估是否进入 Runtime。</li>
      </ul>
    </section>
  </div>

  <footer>
    生成策略：zeusRawLogs-only · min_blob_chars={MIN_BLOB_CHARS}
    · decoding=false · non_zeus_changes={summary["non_zeus_changes"]}
  </footer>
</main>
<script>
  const search = document.getElementById("full-search");
  const rows = Array.from(
    document.querySelectorAll("#full-table tbody tr")
  );
  search.addEventListener("input", () => {{
    const query = search.value.trim().toLowerCase();
    rows.forEach((row) => {{
      row.style.display = row.textContent.toLowerCase().includes(query)
        ? "" : "none";
    }});
  }});
</script>
</body>
</html>
"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_name(f".{output_path.name}.tmp")
    temp_path.write_text(document, encoding="utf-8")
    temp_path.replace(output_path)


def write_markdown(
    output_path: Path,
    *,
    source_path: Path,
    key_frame: pd.DataFrame,
    audit_frame: pd.DataFrame,
    summary: dict[str, Any],
) -> None:
    key_rows = []
    for _, row in key_frame.iterrows():
        key_rows.append(
            "| "
            + " | ".join(
                [
                    str(row["alert_id"]),
                    format_int(row["命中片段数"]),
                    format_int(row["唯一编码值数"]),
                    format_int(row["压缩前字符数"]),
                    format_int(row["压缩后字符数"]),
                    format_int(row["节省字符数"]),
                    format_percent(row["alert_full_data压缩率"]),
                    format_percent(row["zeusRawLogs压缩率"]),
                ]
            )
            + " |"
        )
    kind_frame = kind_summary_frame(audit_frame)
    kind_rows = []
    for _, row in kind_frame.iterrows():
        kind_rows.append(
            "| "
            + " | ".join(
                [
                    str(row["编码类型说明"]),
                    format_int(row["命中片段数"]),
                    format_int(row["唯一编码值数"]),
                    format_int(row["原始字符数"]),
                    format_int(row["节省字符数"]),
                ]
            )
            + " |"
        )

    content = f"""# ZeusRawLogs 长编码压缩技术验证

> 内部敏感材料：配套 HTML 左右对比保留原始 `zeusRawLogs`，可能包含凭证、
> Token、Cookie 或其他敏感内容，仅向有告警访问权限的同事分享。

## 1. 背景与目标

部分告警的 `zeusRawLogs` 含大段 Base64-like、JWT-like 或 Percent-encoded
内容。它们直接进入 LLM 会明显拉长上下文，但多数场景不需要模型读取编码原文。
本验证的目标是在不解码、不修改原始告警的前提下，将这些连续长编码片段替换为
可审计的短占位符。

数据源：`{source_path.relative_to(ROOT)}`。本报告按实际输入 DataFrame 统计，
共 **{summary["rows"]} 条**；文件名中的数量不作为校验口径。

## 2. 处理范围

- 只处理 `alert_full_data` 下任意层级的 `zeusRawLogs`。
- 不处理 `agent_response`，不处理其他字段。
- 不解码，只按编码形态识别；通用 Base64-like/Hex-like 最小长度为
  {MIN_BLOB_CHARS} 字符。
- 命中片段替换为 `<ENCODED:type:length:sha256=short-hash:OMITTED>`。
- 占位符保留 12 位短哈希；审计明细保留 JSON 路径、编码类型、原始长度和完整 SHA-256。
- 原始 pickle 和原始字典不修改，压缩结果只是 LLM 投影验证。

## 3. 总体效果

- 实际样本：**{format_int(summary["rows"])} 条**
- 命中告警：**{format_int(summary["hit_rows"])} 条**
- 命中片段：**{format_int(summary["matches"])} 个**
- 唯一编码值：**{format_int(summary["unique_values"])} 个**
- `alert_full_data`：{format_int(summary["before_chars"])} →
  {format_int(summary["after_chars"])} 字符
- 节省：**{format_int(summary["saved_chars"])} 字符**
- `alert_full_data` 总体字符压缩率：
  **{format_percent(summary["reduction_rate"])}**
- `zeusRawLogs` 自身字符压缩率：
  **{format_percent(summary["zeus_reduction_rate"])}**
- 非 `zeusRawLogs` 字段变化数：**{summary["non_zeus_changes"]}**

注意：以上是紧凑 JSON 的字符数口径，不是特定模型 tokenizer 的精确 token
压缩率。正式评估成本时仍应使用目标模型 tokenizer 复测。

## 4. 命中类型

| 类型 | 命中片段数 | 唯一值数 | 原始字符数 | 节省字符数 |
| --- | ---: | ---: | ---: | ---: |
{chr(10).join(kind_rows)}

## 5. 重点 6 条

| Alert ID | 命中片段 | 唯一值 | 压缩前 | 压缩后 | 节省字符 | alert_full_data 压缩率 | zeusRawLogs 压缩率 |
| ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |
{chr(10).join(key_rows)}

重点样本的完整“压缩了什么”和左右对比见
`zeus_raw_logs_compaction_technical_report.html`。其中 1973909 默认展开，
其余告警可按需展开；Excel 的“重点对比”“字段明细”“命中审计”工作表适合
用于汇报材料取数。

## 6. 为什么 1973909 压缩特别多

该告警不只有当前事件，还包含 `relatedAlertList`。同类大段响应体、
`message`、`payload` 和 JWT-like 内容在当前/相关告警中重复出现，所以总字符
下降明显。命中次数不等于唯一编码值数，Excel 同时给出两者，便于识别重复内容
带来的放大效应。

## 7. 风险和上线前验证

- **误压缩：**长普通文本可能碰巧符合编码字符集。1970506 的
  `http_http_method` 命中是值得复核的语义误判候选。
- **漏压缩：**短编码、分段编码或混杂特殊字符的编码可能未命中。
- **调查信息损失：**LLM 看不到替换后的原文，因此运营界面和审计存储仍需保留
  完整原告警。
- **敏感数据：**为了左右对比，HTML 不脱敏；只可在授权范围内分享。
- **建议验证：**按字段和 Topic 分层抽检误压缩/漏压缩率；使用目标模型
  tokenizer 复测 token 节省；比较压缩前后研判质量、证据引用和人工复核耗时。

本结果是离线技术验证，不代表已经接入 SOC Runtime，也不代表生产可用。
"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_name(f".{output_path.name}.tmp")
    temp_path.write_text(content, encoding="utf-8")
    temp_path.replace(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_DATA_PATH,
        help=f"Input DataFrame pickle (default: {DEFAULT_DATA_PATH})",
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help=f"Excel output (default: {DEFAULT_EXCEL_PATH})",
    )
    parser.add_argument(
        "--html",
        type=Path,
        default=DEFAULT_HTML_PATH,
        help=f"HTML output (default: {DEFAULT_HTML_PATH})",
    )
    parser.add_argument(
        "--markdown",
        type=Path,
        default=DEFAULT_MARKDOWN_PATH,
        help=f"Markdown output (default: {DEFAULT_MARKDOWN_PATH})",
    )
    parser.add_argument(
        "--min-blob-chars",
        type=int,
        default=MIN_BLOB_CHARS,
        help=f"Minimum generic encoded span length (default: {MIN_BLOB_CHARS})",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    frame = load_sample(args.input)
    row_frame, field_frame, audit_frame, key_comparisons = process_sample(
        frame,
        min_blob_chars=args.min_blob_chars,
    )
    missing_key_alerts = [
        alert_id for alert_id in KEY_ALERT_IDS if alert_id not in key_comparisons
    ]
    if missing_key_alerts:
        raise ValueError(f"key alert IDs missing from sample: {missing_key_alerts}")
    key_frame = build_key_frame(row_frame, field_frame)
    summary = summarize(row_frame, audit_frame)

    write_excel(
        args.excel,
        source_path=args.input,
        row_frame=row_frame,
        field_frame=field_frame,
        audit_frame=audit_frame,
        key_frame=key_frame,
        summary=summary,
    )
    write_html(
        args.html,
        source_path=args.input,
        row_frame=row_frame,
        field_frame=field_frame,
        audit_frame=audit_frame,
        key_frame=key_frame,
        key_comparisons=key_comparisons,
        summary=summary,
    )
    write_markdown(
        args.markdown,
        source_path=args.input,
        key_frame=key_frame,
        audit_frame=audit_frame,
        summary=summary,
    )

    print(f"rows={summary['rows']}")
    print(f"hit_rows={summary['hit_rows']}")
    print(f"matches={summary['matches']}")
    print(f"before_chars={summary['before_chars']}")
    print(f"after_chars={summary['after_chars']}")
    print(f"saved_chars={summary['saved_chars']}")
    print(f"reduction_percent={summary['reduction_rate'] * 100:.2f}")
    print(f"zeus_reduction_percent={summary['zeus_reduction_rate'] * 100:.2f}")
    print(f"non_zeus_changes={summary['non_zeus_changes']}")
    print(f"excel={args.excel}")
    print(f"html={args.html}")
    print(f"markdown={args.markdown}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
