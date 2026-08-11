"""Governed, investigation-only access to historical PingAn EDR path data.

The source workbook contains historical ignored-alert samples. It is compiled
once into a versioned SQLite catalog and is never treated as an allowlist.
Path location risk and historical disposition are deliberately separate.
"""

from __future__ import annotations

import hashlib
import json
import ntpath
import os
import re
import sqlite3
from collections import defaultdict
from collections.abc import Iterable, Mapping
from datetime import UTC, datetime
from enum import StrEnum
from pathlib import Path
from typing import Any, Literal
from uuid import uuid4

from pydantic import BaseModel, ConfigDict, Field, field_validator

from soc_agent.actions.adapters import SocActionAdapterRegistryError
from soc_agent.contracts import (
    ServiceRequestContext,
    SocAgentActionAdapterDescriptor,
    SocAgentActionCommand,
    SocAgentActionResult,
    SocAgentRiskLevel,
)

PINGAN_SOFTWARE_PATH_LOOKUP_ACTION = "endpoint.software_path.lookup"
PINGAN_SOFTWARE_PATH_CATALOG_SCHEMA = "soc.pingan_software_path_catalog.v2"
PINGAN_SOFTWARE_PATH_BUILD_REPORT_SCHEMA = "soc.pingan_software_path_catalog_build.v2"
PINGAN_SOFTWARE_PATH_RESULT_SCHEMA = "soc.pingan_software_path_context.v2"
DEFAULT_CATALOG_ENV = "SOC_PINGAN_SOFTWARE_PATH_CATALOG_PATH"

_REQUIRED_WORKBOOK_COLUMNS = frozenset({"alertId", "flag", "zeusRawLogs", "path_parser"})
_EXECUTABLE_EXTENSIONS = frozenset({".bat", ".cmd", ".com", ".dll", ".exe", ".jar", ".js", ".msi", ".ps1", ".scr", ".sys", ".vbs"})
_MD5_PATTERN = re.compile(r"^[0-9a-f]{32}$")
_PATH_FAMILY_PLACEHOLDER = "{dynamic_segment}"
_MIN_PATH_FAMILY_MEMBERS = 2
_DEPLOYMENT_CACHE_PARENT_NAMES = frozenset({"ccmcache"})


class PingAnSoftwarePathCatalogError(RuntimeError):
    """Base error for catalog compilation and lookup failures."""


class PingAnSoftwarePathControlZone(StrEnum):
    """Tenant path-control context; it is not a maliciousness verdict."""

    MANAGED_SYSTEM = "managed_system"
    MANAGED_PROGRAM = "managed_program"
    LESS_MANAGED = "less_managed"
    USER_WRITABLE = "user_writable"
    TEMPORARY = "temporary"
    DEVICE_OR_UNC = "device_or_unc"
    UNKNOWN = "unknown"


class PingAnSoftwarePathAttention(StrEnum):
    LOWER = "lower"
    NORMAL = "normal"
    HIGH = "high"


class PingAnSoftwarePathMatchType(StrEnum):
    NONE = "none"
    EXACT_PATH = "exact_path"
    EXACT_PATH_AND_MD5 = "exact_path_and_md5"
    EXACT_PATH_HASH_MISMATCH = "exact_path_hash_mismatch"
    PATH_FAMILY = "path_family"
    PATH_FAMILY_AND_MD5 = "path_family_and_md5"
    PATH_FAMILY_HASH_MISMATCH = "path_family_hash_mismatch"


class PingAnSoftwarePathFreshness(StrEnum):
    CURRENT = "current"
    STALE = "stale"
    UNKNOWN = "unknown"


class _StrictModel(BaseModel):
    model_config = ConfigDict(extra="forbid")


class PingAnSoftwarePathCatalogBuildReport(_StrictModel):
    schema_version: Literal["soc.pingan_software_path_catalog_build.v2"] = PINGAN_SOFTWARE_PATH_BUILD_REPORT_SCHEMA
    catalog_schema_version: Literal["soc.pingan_software_path_catalog.v2"] = PINGAN_SOFTWARE_PATH_CATALOG_SCHEMA
    catalog_id: str
    catalog_path: str
    source_path: str
    source_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    generated_at: datetime
    row_count: int = Field(ge=0)
    parsed_row_count: int = Field(ge=0)
    malformed_path_parser_count: int = Field(ge=0)
    malformed_raw_log_count: int = Field(ge=0)
    path_entry_count: int = Field(ge=0)
    observation_count: int = Field(ge=0)
    executable_like_entry_count: int = Field(ge=0)
    path_family_count: int = Field(ge=0)
    path_family_member_count: int = Field(ge=0)
    control_zone_counts: dict[str, int] = Field(default_factory=dict)
    legacy_bucket_counts: dict[str, int] = Field(default_factory=dict)
    source_flag_counts: dict[str, int] = Field(default_factory=dict)
    candidate_only: Literal[True] = True
    allowlist: Literal[False] = False
    evidence_boundary: Literal["investigation_only"] = "investigation_only"
    decision_impact: Literal["none"] = "none"


class PingAnSoftwarePathHistoricalContext(_StrictModel):
    candidate_status: Literal["historical_candidate"] = "historical_candidate"
    source_dispositions: list[str] = Field(default_factory=list)
    legacy_path_buckets: list[str] = Field(default_factory=list)
    occurrence_count: int = Field(ge=0)
    source_alert_count: int = Field(ge=0)
    first_seen_at: datetime | None = None
    last_seen_at: datetime | None = None
    freshness: PingAnSoftwarePathFreshness = PingAnSoftwarePathFreshness.UNKNOWN
    process_names: list[str] = Field(default_factory=list, max_length=20)
    known_md5s: list[str] = Field(default_factory=list, max_length=20)
    rule_codes: list[str] = Field(default_factory=list, max_length=20)
    source_alert_ids: list[str] = Field(default_factory=list, max_length=10)
    source_alert_ids_truncated: bool = False


class PingAnSoftwarePathFamilyContext(_StrictModel):
    """One conservatively inferred family built only from legacy safe_paths."""

    family_id: str = Field(min_length=1, max_length=128)
    pattern_path: str = Field(min_length=1, max_length=4096)
    variable_segment_kind: str = Field(min_length=1, max_length=64)
    query_variable_segment: str = Field(min_length=1, max_length=512)
    member_path_count: int = Field(ge=2)
    source_alert_count: int = Field(ge=1)
    occurrence_count: int = Field(ge=1)
    known_md5s: list[str] = Field(default_factory=list, max_length=20)
    source_alert_ids: list[str] = Field(default_factory=list, max_length=10)
    source_alert_ids_truncated: bool = False


class PingAnSoftwarePathContextResult(_StrictModel):
    schema_version: Literal["soc.pingan_software_path_context.v2"] = PINGAN_SOFTWARE_PATH_RESULT_SCHEMA
    query_path: str = Field(min_length=1, max_length=4096)
    normalized_path: str = Field(min_length=1, max_length=4096)
    query_md5: str | None = None
    matched: bool = False
    match_type: PingAnSoftwarePathMatchType = PingAnSoftwarePathMatchType.NONE
    control_zone: PingAnSoftwarePathControlZone
    location_attention: PingAnSoftwarePathAttention
    historical_context: PingAnSoftwarePathHistoricalContext | None = None
    path_family_context: PingAnSoftwarePathFamilyContext | None = None
    exact_safe_path_candidate: bool = False
    catalog_id: str
    catalog_schema_version: Literal["soc.pingan_software_path_catalog.v2"] = PINGAN_SOFTWARE_PATH_CATALOG_SCHEMA
    source_sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    warnings: list[str] = Field(default_factory=list)
    provider_mode: Literal["local_catalog"] = "local_catalog"
    mocked: Literal[False] = False
    candidate_only: Literal[True] = True
    allowlist: Literal[False] = False
    evidence_boundary: Literal["investigation_only"] = "investigation_only"
    decision_impact: Literal["none"] = "none"
    automation_eligible: Literal[False] = False
    raw_rows_included: Literal[False] = False

    @field_validator("query_md5")
    @classmethod
    def validate_md5(cls, value: str | None) -> str | None:
        if value is None:
            return None
        normalized = value.strip().lower()
        if not _MD5_PATTERN.fullmatch(normalized):
            raise ValueError("query_md5 must be a 32-character hexadecimal MD5")
        return normalized


class PingAnSoftwarePathCatalog:
    """Read-only exact and inferred-family lookup over a compiled local catalog."""

    def __init__(self, path: str | Path, *, freshness_days: int = 180) -> None:
        self.path = Path(path).expanduser().resolve()
        if freshness_days < 1:
            raise ValueError("freshness_days must be positive")
        self.freshness_days = freshness_days
        if not self.path.is_file():
            raise PingAnSoftwarePathCatalogError(f"software path catalog does not exist: {self.path}")
        metadata = self._metadata()
        if metadata.get("schema_version") != PINGAN_SOFTWARE_PATH_CATALOG_SCHEMA:
            raise PingAnSoftwarePathCatalogError("unsupported PingAn software path catalog schema")
        self.catalog_id = _required_metadata(metadata, "catalog_id")
        self.source_sha256 = _required_metadata(metadata, "source_sha256")

    @classmethod
    def from_env(cls, environ: Mapping[str, str] | None = None) -> PingAnSoftwarePathCatalog:
        values = os.environ if environ is None else environ
        configured = values.get(DEFAULT_CATALOG_ENV, "").strip()
        if not configured:
            raise PingAnSoftwarePathCatalogError(f"{DEFAULT_CATALOG_ENV} is required")
        freshness_days = int(values.get("SOC_PINGAN_SOFTWARE_PATH_FRESHNESS_DAYS", "180"))
        return cls(configured, freshness_days=freshness_days)

    def lookup(
        self,
        path: str,
        *,
        md5: str | None = None,
        as_of: datetime | None = None,
    ) -> PingAnSoftwarePathContextResult:
        normalized_path = normalize_windows_path(path)
        normalized_md5 = _normalize_optional_md5(md5)
        control_zone, attention = classify_pingan_path(normalized_path)
        row = self._entry(normalized_path)
        family_match = self._family_entry(normalized_path)
        warnings = _path_warnings(control_zone)
        if row is None and family_match is None:
            return PingAnSoftwarePathContextResult(
                query_path=path,
                normalized_path=normalized_path,
                query_md5=normalized_md5,
                control_zone=control_zone,
                location_attention=attention,
                catalog_id=self.catalog_id,
                source_sha256=self.source_sha256,
                warnings=warnings,
            )

        family_context = None
        family_match_type = None
        if family_match is not None:
            family_row, query_variable_segment = family_match
            family_known_md5s = _json_string_list(family_row["md5s_json"], limit=20)
            if normalized_md5 and family_known_md5s and normalized_md5 not in family_known_md5s:
                family_match_type = PingAnSoftwarePathMatchType.PATH_FAMILY_HASH_MISMATCH
            elif normalized_md5 and normalized_md5 in family_known_md5s:
                family_match_type = PingAnSoftwarePathMatchType.PATH_FAMILY_AND_MD5
            else:
                family_match_type = PingAnSoftwarePathMatchType.PATH_FAMILY
            family_alert_ids = _json_string_list(family_row["source_alert_ids_json"], limit=11)
            family_context = PingAnSoftwarePathFamilyContext(
                family_id=str(family_row["family_id"]),
                pattern_path=str(family_row["pattern_path"]),
                variable_segment_kind=str(family_row["variable_segment_kind"]),
                query_variable_segment=query_variable_segment,
                member_path_count=int(family_row["member_path_count"]),
                source_alert_count=int(family_row["source_alert_count"]),
                occurrence_count=int(family_row["occurrence_count"]),
                known_md5s=family_known_md5s,
                source_alert_ids=family_alert_ids[:10],
                source_alert_ids_truncated=len(family_alert_ids) > 10,
            )
            warnings.append("Path matched an inferred family compiled only from historical safe_paths; the family pattern and member lineage are retained for audit.")
            if family_match_type is PingAnSoftwarePathMatchType.PATH_FAMILY_HASH_MISMATCH:
                warnings.append("The path family matched, but the supplied MD5 differs from every cataloged family MD5.")

        if row is None:
            assert family_match_type is not None
            return PingAnSoftwarePathContextResult(
                query_path=path,
                normalized_path=normalized_path,
                query_md5=normalized_md5,
                matched=True,
                match_type=family_match_type,
                control_zone=control_zone,
                location_attention=attention,
                path_family_context=family_context,
                catalog_id=self.catalog_id,
                source_sha256=self.source_sha256,
                warnings=warnings,
            )

        known_md5s = _json_string_list(row["md5s_json"], limit=20)
        if normalized_md5 and known_md5s and normalized_md5 not in known_md5s:
            match_type = PingAnSoftwarePathMatchType.EXACT_PATH_HASH_MISMATCH
            warnings.append("The exact path was seen historically, but the supplied MD5 differs from every cataloged MD5.")
        elif normalized_md5 and normalized_md5 in known_md5s:
            match_type = PingAnSoftwarePathMatchType.EXACT_PATH_AND_MD5
        else:
            match_type = PingAnSoftwarePathMatchType.EXACT_PATH

        last_seen = _parse_datetime(row["last_seen_at"])
        freshness = _freshness(last_seen, as_of=as_of, freshness_days=self.freshness_days)
        if freshness is PingAnSoftwarePathFreshness.STALE:
            warnings.append("Historical path context is stale and requires current behavior or ownership confirmation.")
        warnings.append("Historical ignored disposition is context only; it is not an allowlist or benign verdict.")
        alert_ids = self._source_alert_ids(row["path_hash"], limit=11)
        historical = PingAnSoftwarePathHistoricalContext(
            source_dispositions=_json_string_list(row["source_dispositions_json"]),
            legacy_path_buckets=_json_string_list(row["legacy_buckets_json"]),
            occurrence_count=int(row["occurrence_count"]),
            source_alert_count=int(row["source_alert_count"]),
            first_seen_at=_parse_datetime(row["first_seen_at"]),
            last_seen_at=last_seen,
            freshness=freshness,
            process_names=_json_string_list(row["process_names_json"], limit=20),
            known_md5s=known_md5s,
            rule_codes=_json_string_list(row["rule_codes_json"], limit=20),
            source_alert_ids=alert_ids[:10],
            source_alert_ids_truncated=len(alert_ids) > 10,
        )
        return PingAnSoftwarePathContextResult(
            query_path=path,
            normalized_path=normalized_path,
            query_md5=normalized_md5,
            matched=True,
            match_type=match_type,
            control_zone=control_zone,
            location_attention=attention,
            historical_context=historical,
            path_family_context=family_context,
            exact_safe_path_candidate="safe_paths" in historical.legacy_path_buckets,
            catalog_id=self.catalog_id,
            source_sha256=self.source_sha256,
            warnings=warnings,
        )

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(f"file:{self.path}?mode=ro", uri=True)
        connection.row_factory = sqlite3.Row
        return connection

    def _metadata(self) -> dict[str, str]:
        try:
            with self._connect() as connection:
                rows = connection.execute("SELECT key, value FROM catalog_metadata").fetchall()
        except sqlite3.Error as exc:
            raise PingAnSoftwarePathCatalogError("cannot read PingAn software path catalog metadata") from exc
        return {str(row["key"]): str(row["value"]) for row in rows}

    def _entry(self, normalized_path: str) -> sqlite3.Row | None:
        with self._connect() as connection:
            return connection.execute(
                "SELECT * FROM path_entries WHERE path_hash = ? AND normalized_path = ?",
                (_sha256_text(normalized_path), normalized_path),
            ).fetchone()

    def _family_entry(self, normalized_path: str) -> tuple[sqlite3.Row, str] | None:
        components = _path_components(normalized_path)
        if len(components) < 3:
            return None
        file_name = components[-1]
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT * FROM path_families
                WHERE file_name = ? AND segment_count = ?
                ORDER BY member_path_count DESC, family_id
                """,
                (file_name, len(components)),
            ).fetchall()
        for row in rows:
            pattern_components = _path_components(str(row["pattern_path"]))
            if len(pattern_components) != len(components):
                continue
            variable_index = int(row["variable_index"])
            if all(index == variable_index or actual == expected for index, (actual, expected) in enumerate(zip(components, pattern_components, strict=True))):
                return row, components[variable_index]
        return None

    def _source_alert_ids(self, path_hash: str, *, limit: int) -> list[str]:
        with self._connect() as connection:
            rows = connection.execute(
                "SELECT DISTINCT alert_id FROM path_observations WHERE path_hash = ? ORDER BY alert_id LIMIT ?",
                (path_hash, limit),
            ).fetchall()
        return [str(row["alert_id"]) for row in rows]


class PingAnSoftwarePathLookupActionAdapter:
    """Read-only action adapter over the local historical catalog."""

    def __init__(
        self,
        catalog: PingAnSoftwarePathCatalog,
        *,
        descriptor: SocAgentActionAdapterDescriptor | None = None,
    ) -> None:
        self.catalog = catalog
        self.descriptor = descriptor or pingan_software_path_adapter_descriptor()
        if self.descriptor.route != PINGAN_SOFTWARE_PATH_LOOKUP_ACTION or self.descriptor.action != PINGAN_SOFTWARE_PATH_LOOKUP_ACTION:
            raise SocActionAdapterRegistryError("PingAnSoftwarePathLookupActionAdapter requires route/action endpoint.software_path.lookup")
        if self.descriptor.risk_level is not SocAgentRiskLevel.READ_ONLY or self.descriptor.external_side_effect != "read":
            raise SocActionAdapterRegistryError("PingAn software path lookup must be a read-only adapter")

    def dry_run(
        self,
        command: SocAgentActionCommand,
        *,
        context: ServiceRequestContext,
    ) -> SocAgentActionResult:
        self._validate(command, dry_run=True)
        path = _required_payload_string(command.payload, "path")
        return SocAgentActionResult(
            route=command.route,
            action=command.action,
            status="success",
            message="Software path lookup dry-run validated; the local catalog was not read.",
            payload={
                "adapter_id": self.descriptor.adapter_id,
                "dry_run": True,
                "path_hash": _sha256_text(normalize_windows_path(path)),
                "external_side_effect": "not_executed",
                "evidence_boundary": "investigation_only",
                "decision_impact": "none",
                "automation_eligible": False,
                "executed_by": context.actor.model_dump(mode="json"),
            },
        )

    def execute(
        self,
        command: SocAgentActionCommand,
        *,
        context: ServiceRequestContext,
    ) -> SocAgentActionResult:
        self._validate(command, dry_run=False)
        path = _required_payload_string(command.payload, "path")
        md5 = command.payload.get("md5")
        result = self.catalog.lookup(path, md5=md5 if isinstance(md5, str) else None)
        return SocAgentActionResult(
            route=command.route,
            action=command.action,
            status="success",
            message=("Historical software path context found; behavior verification remains required." if result.matched else "No historical software path context matched; location risk remains available."),
            payload={
                "adapter_id": self.descriptor.adapter_id,
                "adapter_kind": self.descriptor.adapter_kind,
                "dry_run": False,
                **result.model_dump(mode="json", exclude_none=True),
                "executed_by": context.actor.model_dump(mode="json"),
            },
        )

    def _validate(self, command: SocAgentActionCommand, *, dry_run: bool) -> None:
        if command.route != self.descriptor.route or command.action != self.descriptor.action:
            raise SocActionAdapterRegistryError("software path command does not match its adapter descriptor")
        if command.dry_run is not dry_run:
            expected = "true" if dry_run else "false"
            raise SocActionAdapterRegistryError(f"software path command requires dry_run={expected}")
        _required_payload_string(command.payload, "path")
        md5 = command.payload.get("md5")
        if md5 is not None:
            _normalize_optional_md5(str(md5))


def pingan_software_path_adapter_descriptor(
    *,
    adapter_id: str = "pingan-software-path-local-catalog",
) -> SocAgentActionAdapterDescriptor:
    return SocAgentActionAdapterDescriptor(
        adapter_id=adapter_id,
        route=PINGAN_SOFTWARE_PATH_LOOKUP_ACTION,
        action=PINGAN_SOFTWARE_PATH_LOOKUP_ACTION,
        risk_level=SocAgentRiskLevel.READ_ONLY,
        adapter_kind="service",
        external_side_effect="read",
        dry_run_supported=True,
        execute_supported=True,
        idempotency_required=False,
        required_payload_fields=["path"],
        description="Read-only lookup of historical PingAn EDR software path context.",
        metadata={
            "tenant": "pingan",
            "candidate_only": True,
            "allowlist": False,
            "evidence_boundary": "investigation_only",
            "decision_impact": "none",
        },
    )


def compile_pingan_software_path_catalog(
    workbook_path: str | Path,
    catalog_path: str | Path,
) -> PingAnSoftwarePathCatalogBuildReport:
    """Compile the historical workbook into an atomically replaced SQLite catalog."""

    source = Path(workbook_path).expanduser().resolve()
    target = Path(catalog_path).expanduser().resolve()
    if not source.is_file():
        raise PingAnSoftwarePathCatalogError(f"source workbook does not exist: {source}")
    source_sha256 = _sha256_file(source)
    generated_at = datetime.now(UTC)
    catalog_id = f"PSC-{source_sha256[:16].upper()}"
    entries, observations, metrics = _read_workbook(source)
    path_families = _build_path_families(entries, observations)
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f".{target.name}.{uuid4().hex}.tmp")
    try:
        _write_catalog(
            temporary,
            entries=entries,
            observations=observations,
            path_families=path_families,
            metadata={
                "schema_version": PINGAN_SOFTWARE_PATH_CATALOG_SCHEMA,
                "catalog_id": catalog_id,
                "source_sha256": source_sha256,
                "source_filename": source.name,
                "generated_at": generated_at.isoformat(),
                "candidate_only": "true",
                "allowlist": "false",
                "evidence_boundary": "investigation_only",
                "decision_impact": "none",
                "path_family_inference": "safe_paths_only",
            },
        )
        os.replace(temporary, target)
        target.chmod(0o600)
    finally:
        temporary.unlink(missing_ok=True)

    return PingAnSoftwarePathCatalogBuildReport(
        catalog_id=catalog_id,
        catalog_path=str(target),
        source_path=str(source),
        source_sha256=source_sha256,
        generated_at=generated_at,
        row_count=metrics["row_count"],
        parsed_row_count=metrics["parsed_row_count"],
        malformed_path_parser_count=metrics["malformed_path_parser_count"],
        malformed_raw_log_count=metrics["malformed_raw_log_count"],
        path_entry_count=len(entries),
        observation_count=len(observations),
        executable_like_entry_count=sum(bool(entry["executable_like"]) for entry in entries.values()),
        path_family_count=len(path_families),
        path_family_member_count=sum(len(family["member_path_hashes"]) for family in path_families.values()),
        control_zone_counts=_count_values(entry["control_zone"] for entry in entries.values()),
        legacy_bucket_counts=_count_values(observation["legacy_bucket"] for observation in observations),
        source_flag_counts=metrics["source_flag_counts"],
    )


def normalize_windows_path(value: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError("path must be a non-empty string")
    text = re.sub(r"\s+", " ", value.strip().strip('"')).replace("/", "\\")
    if text.startswith("\\\\?\\"):
        text = text[4:]
    is_unc = text.startswith("\\\\")
    normalized = ntpath.normpath(text)
    if is_unc and not normalized.startswith("\\\\"):
        normalized = "\\" + normalized
    return normalized.lower()


def is_executable_like_path(value: str) -> bool:
    try:
        normalized = normalize_windows_path(value)
    except ValueError:
        return False
    return ntpath.splitext(normalized)[1].lower() in _EXECUTABLE_EXTENSIONS


def _path_components(value: str) -> list[str]:
    normalized = normalize_windows_path(value)
    drive, tail = ntpath.splitdrive(normalized)
    segments = [segment for segment in tail.strip("\\").split("\\") if segment]
    return [drive.lower(), *segments] if drive else segments


def _join_path_components(components: list[str]) -> str:
    if not components:
        raise ValueError("path family components cannot be empty")
    head, *tail = components
    return head + "\\" + "\\".join(tail) if tail else head


def _variable_segment_kind(segment: str, *, parent: str) -> str | None:
    normalized = segment.casefold()
    if parent.casefold() in _DEPLOYMENT_CACHE_PARENT_NAMES and re.fullmatch(r"[a-z0-9]{1,4}", normalized):
        return "deployment_cache_slot"
    if re.fullmatch(r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}", normalized):
        return "uuid"
    if len(normalized) >= 4 and normalized.isdigit():
        return "numeric"
    if len(normalized) >= 8 and re.fullmatch(r"[0-9a-f]+", normalized):
        return "hex"
    return None


def classify_pingan_path(
    normalized_path: str,
) -> tuple[PingAnSoftwarePathControlZone, PingAnSoftwarePathAttention]:
    path = normalize_windows_path(normalized_path)
    if path.startswith("\\device\\") or path.startswith("\\\\"):
        return PingAnSoftwarePathControlZone.DEVICE_OR_UNC, PingAnSoftwarePathAttention.NORMAL
    if _is_temporary_path(path):
        return PingAnSoftwarePathControlZone.TEMPORARY, PingAnSoftwarePathAttention.HIGH
    if _is_user_writable_path(path):
        return PingAnSoftwarePathControlZone.USER_WRITABLE, PingAnSoftwarePathAttention.HIGH
    if path.startswith("d:\\"):
        return PingAnSoftwarePathControlZone.LESS_MANAGED, PingAnSoftwarePathAttention.HIGH
    if path.startswith(("c:\\windows\\", "c:\\recovery\\", "c:\\boot\\", "c:\\perflogs\\")):
        return PingAnSoftwarePathControlZone.MANAGED_SYSTEM, PingAnSoftwarePathAttention.LOWER
    if path.startswith(("c:\\program files\\", "c:\\program files (x86)\\")):
        return PingAnSoftwarePathControlZone.MANAGED_PROGRAM, PingAnSoftwarePathAttention.LOWER
    return PingAnSoftwarePathControlZone.UNKNOWN, PingAnSoftwarePathAttention.NORMAL


def _read_workbook(
    source: Path,
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - locked DeerFlow install includes xlsx support
        raise PingAnSoftwarePathCatalogError("openpyxl is required to compile the PingAn path workbook") from exc

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration as exc:
            raise PingAnSoftwarePathCatalogError("source workbook is empty") from exc
        headers = [str(value).strip() if value is not None else "" for value in header_values]
        missing = _REQUIRED_WORKBOOK_COLUMNS.difference(headers)
        if missing:
            raise PingAnSoftwarePathCatalogError(f"source workbook is missing columns: {sorted(missing)}")
        indexes = {name: headers.index(name) for name in _REQUIRED_WORKBOOK_COLUMNS}
        entries: dict[str, dict[str, Any]] = {}
        observations: list[dict[str, Any]] = []
        metrics: dict[str, Any] = {
            "row_count": 0,
            "parsed_row_count": 0,
            "malformed_path_parser_count": 0,
            "malformed_raw_log_count": 0,
            "source_flag_counts": defaultdict(int),
        }
        for source_row, values in enumerate(rows, start=2):
            metrics["row_count"] += 1
            alert_id = _cell_text(values[indexes["alertId"]]) or f"row-{source_row}"
            source_flag = _cell_text(values[indexes["flag"]]) or "unknown"
            metrics["source_flag_counts"][source_flag] += 1
            parsed_paths = _parse_json_object(values[indexes["path_parser"]])
            if parsed_paths is None:
                metrics["malformed_path_parser_count"] += 1
                continue
            path_groups = parsed_paths.get("paths")
            if not isinstance(path_groups, Mapping):
                metrics["malformed_path_parser_count"] += 1
                continue
            logs = _parse_json_list(values[indexes["zeusRawLogs"]])
            if logs is None:
                metrics["malformed_raw_log_count"] += 1
                logs = []
            path_metadata, row_times = _path_metadata(logs)
            row_seen: set[tuple[str, str]] = set()
            row_had_path = False
            for bucket, raw_paths in (
                ("safe_paths", path_groups.get("safe_paths")),
                ("other_paths", path_groups.get("other_paths")),
            ):
                if not isinstance(raw_paths, list):
                    continue
                for raw_path in raw_paths:
                    if not isinstance(raw_path, str) or not raw_path.strip():
                        continue
                    try:
                        normalized_path = normalize_windows_path(raw_path)
                    except ValueError:
                        continue
                    if (normalized_path, bucket) in row_seen:
                        continue
                    row_seen.add((normalized_path, bucket))
                    row_had_path = True
                    control_zone, attention = classify_pingan_path(normalized_path)
                    path_hash = _sha256_text(normalized_path)
                    metadata = path_metadata.get(normalized_path, {})
                    observation = {
                        "observation_id": _sha256_text(f"{source_row}|{alert_id}|{bucket}|{normalized_path}"),
                        "path_hash": path_hash,
                        "alert_id": alert_id,
                        "source_row": source_row,
                        "legacy_bucket": bucket,
                        "source_disposition": source_flag,
                        "observed_at": metadata.get("observed_at") or row_times.get("last_seen_at"),
                        "process_name": metadata.get("process_name"),
                        "md5": metadata.get("md5"),
                        "rule_code": metadata.get("rule_code"),
                    }
                    observations.append(observation)
                    entry = entries.setdefault(
                        path_hash,
                        {
                            "path_hash": path_hash,
                            "normalized_path": normalized_path,
                            "display_path": raw_path.strip(),
                            "drive": _drive(normalized_path),
                            "control_zone": control_zone.value,
                            "location_attention": attention.value,
                            "executable_like": int(ntpath.splitext(normalized_path)[1].lower() in _EXECUTABLE_EXTENSIONS),
                            "occurrence_count": 0,
                            "alert_ids": set(),
                            "first_seen_at": None,
                            "last_seen_at": None,
                            "legacy_buckets": set(),
                            "source_dispositions": set(),
                            "process_names": set(),
                            "md5s": set(),
                            "rule_codes": set(),
                        },
                    )
                    entry["occurrence_count"] += 1
                    entry["alert_ids"].add(alert_id)
                    entry["legacy_buckets"].add(bucket)
                    entry["source_dispositions"].add(source_flag)
                    _add_if_text(entry["process_names"], observation["process_name"])
                    _add_if_text(entry["md5s"], observation["md5"])
                    _add_if_text(entry["rule_codes"], observation["rule_code"])
                    _update_seen_range(entry, observation["observed_at"])
            if row_had_path:
                metrics["parsed_row_count"] += 1
        metrics["source_flag_counts"] = dict(sorted(metrics["source_flag_counts"].items()))
        return entries, observations, metrics
    finally:
        workbook.close()


def _build_path_families(
    entries: Mapping[str, Mapping[str, Any]],
    observations: Iterable[Mapping[str, Any]],
) -> dict[str, dict[str, Any]]:
    """Infer one-variable path families only from executable legacy safe_paths."""

    safe_observations: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for observation in observations:
        if observation["legacy_bucket"] == "safe_paths":
            safe_observations[str(observation["path_hash"])].append(observation)

    candidates: dict[tuple[str, str], dict[str, Any]] = {}
    for path_hash, entry in entries.items():
        if not entry["executable_like"] or path_hash not in safe_observations:
            continue
        components = _path_components(str(entry["normalized_path"]))
        if len(components) < 4:
            continue
        for variable_index in range(2, len(components) - 1):
            variable_segment = components[variable_index]
            variable_kind = _variable_segment_kind(
                variable_segment,
                parent=components[variable_index - 1],
            )
            if variable_kind is None:
                continue
            pattern_components = list(components)
            pattern_components[variable_index] = _PATH_FAMILY_PLACEHOLDER
            pattern_path = _join_path_components(pattern_components)
            group = candidates.setdefault(
                (pattern_path, variable_kind),
                {
                    "pattern_path": pattern_path,
                    "file_name": components[-1],
                    "segment_count": len(components),
                    "variable_index": variable_index,
                    "variable_segment_kind": variable_kind,
                    "variable_segments": set(),
                    "member_path_hashes": set(),
                },
            )
            group["variable_segments"].add(variable_segment)
            group["member_path_hashes"].add(path_hash)

    families: dict[str, dict[str, Any]] = {}
    for group in candidates.values():
        if len(group["member_path_hashes"]) < _MIN_PATH_FAMILY_MEMBERS or len(group["variable_segments"]) < _MIN_PATH_FAMILY_MEMBERS:
            continue
        member_observations = [observation for path_hash in group["member_path_hashes"] for observation in safe_observations[path_hash]]
        source_alert_ids = {str(item["alert_id"]) for item in member_observations}
        if len(source_alert_ids) < _MIN_PATH_FAMILY_MEMBERS:
            continue
        family_hash = _sha256_text(group["pattern_path"])
        families[family_hash] = {
            **group,
            "family_id": f"PSF-{family_hash[:20].upper()}",
            "source_alert_ids": source_alert_ids,
            "occurrence_count": len(member_observations),
            "first_seen_at": _minimum_text(item.get("observed_at") for item in member_observations),
            "last_seen_at": _maximum_text(item.get("observed_at") for item in member_observations),
            "process_names": _non_empty_text_set(item.get("process_name") for item in member_observations),
            "md5s": _non_empty_text_set(item.get("md5") for item in member_observations),
            "rule_codes": _non_empty_text_set(item.get("rule_code") for item in member_observations),
        }
    return families


def _write_catalog(
    path: Path,
    *,
    entries: Mapping[str, Mapping[str, Any]],
    observations: Iterable[Mapping[str, Any]],
    path_families: Mapping[str, Mapping[str, Any]],
    metadata: Mapping[str, str],
) -> None:
    with sqlite3.connect(path) as connection:
        connection.executescript(
            """
            PRAGMA journal_mode=DELETE;
            PRAGMA synchronous=FULL;
            CREATE TABLE catalog_metadata (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            CREATE TABLE path_entries (
                path_hash TEXT PRIMARY KEY,
                normalized_path TEXT NOT NULL UNIQUE,
                display_path TEXT NOT NULL,
                drive TEXT,
                control_zone TEXT NOT NULL,
                location_attention TEXT NOT NULL,
                executable_like INTEGER NOT NULL,
                occurrence_count INTEGER NOT NULL,
                source_alert_count INTEGER NOT NULL,
                first_seen_at TEXT,
                last_seen_at TEXT,
                legacy_buckets_json TEXT NOT NULL,
                source_dispositions_json TEXT NOT NULL,
                process_names_json TEXT NOT NULL,
                md5s_json TEXT NOT NULL,
                rule_codes_json TEXT NOT NULL
            );
            CREATE TABLE path_observations (
                observation_id TEXT PRIMARY KEY,
                path_hash TEXT NOT NULL,
                alert_id TEXT NOT NULL,
                source_row INTEGER NOT NULL,
                legacy_bucket TEXT NOT NULL,
                source_disposition TEXT NOT NULL,
                observed_at TEXT,
                process_name TEXT,
                md5 TEXT,
                rule_code TEXT,
                FOREIGN KEY(path_hash) REFERENCES path_entries(path_hash)
            );
            CREATE INDEX ix_path_observations_path_hash ON path_observations(path_hash);
            CREATE INDEX ix_path_observations_alert_id ON path_observations(alert_id);
            CREATE TABLE path_families (
                family_id TEXT PRIMARY KEY,
                pattern_path TEXT NOT NULL UNIQUE,
                file_name TEXT NOT NULL,
                segment_count INTEGER NOT NULL,
                variable_index INTEGER NOT NULL,
                variable_segment_kind TEXT NOT NULL,
                member_path_count INTEGER NOT NULL,
                source_alert_count INTEGER NOT NULL,
                occurrence_count INTEGER NOT NULL,
                first_seen_at TEXT,
                last_seen_at TEXT,
                process_names_json TEXT NOT NULL,
                md5s_json TEXT NOT NULL,
                rule_codes_json TEXT NOT NULL,
                source_alert_ids_json TEXT NOT NULL
            );
            CREATE TABLE path_family_members (
                family_id TEXT NOT NULL,
                path_hash TEXT NOT NULL,
                PRIMARY KEY(family_id, path_hash),
                FOREIGN KEY(family_id) REFERENCES path_families(family_id),
                FOREIGN KEY(path_hash) REFERENCES path_entries(path_hash)
            );
            CREATE INDEX ix_path_families_file_segments ON path_families(file_name, segment_count);
            CREATE INDEX ix_path_family_members_path_hash ON path_family_members(path_hash);
            """
        )
        connection.executemany(
            "INSERT INTO catalog_metadata(key, value) VALUES (?, ?)",
            sorted(metadata.items()),
        )
        connection.executemany(
            """
            INSERT INTO path_entries(
                path_hash, normalized_path, display_path, drive, control_zone,
                location_attention, executable_like, occurrence_count,
                source_alert_count, first_seen_at, last_seen_at,
                legacy_buckets_json, source_dispositions_json,
                process_names_json, md5s_json, rule_codes_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    entry["path_hash"],
                    entry["normalized_path"],
                    entry["display_path"],
                    entry["drive"],
                    entry["control_zone"],
                    entry["location_attention"],
                    entry["executable_like"],
                    entry["occurrence_count"],
                    len(entry["alert_ids"]),
                    entry["first_seen_at"],
                    entry["last_seen_at"],
                    _json_set(entry["legacy_buckets"]),
                    _json_set(entry["source_dispositions"]),
                    _json_set(entry["process_names"], limit=100),
                    _json_set(entry["md5s"], limit=100),
                    _json_set(entry["rule_codes"], limit=100),
                )
                for entry in sorted(entries.values(), key=lambda item: item["normalized_path"])
            ],
        )
        connection.executemany(
            """
            INSERT INTO path_observations(
                observation_id, path_hash, alert_id, source_row, legacy_bucket,
                source_disposition, observed_at, process_name, md5, rule_code
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    item["observation_id"],
                    item["path_hash"],
                    item["alert_id"],
                    item["source_row"],
                    item["legacy_bucket"],
                    item["source_disposition"],
                    item["observed_at"],
                    item["process_name"],
                    item["md5"],
                    item["rule_code"],
                )
                for item in observations
            ],
        )
        connection.executemany(
            """
            INSERT INTO path_families(
                family_id, pattern_path, file_name, segment_count,
                variable_index, variable_segment_kind, member_path_count,
                source_alert_count, occurrence_count, first_seen_at,
                last_seen_at, process_names_json, md5s_json,
                rule_codes_json, source_alert_ids_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    family["family_id"],
                    family["pattern_path"],
                    family["file_name"],
                    family["segment_count"],
                    family["variable_index"],
                    family["variable_segment_kind"],
                    len(family["member_path_hashes"]),
                    len(family["source_alert_ids"]),
                    family["occurrence_count"],
                    family["first_seen_at"],
                    family["last_seen_at"],
                    _json_set(family["process_names"], limit=100),
                    _json_set(family["md5s"], limit=100),
                    _json_set(family["rule_codes"], limit=100),
                    _json_set(family["source_alert_ids"]),
                )
                for family in sorted(path_families.values(), key=lambda item: item["pattern_path"])
            ],
        )
        connection.executemany(
            "INSERT INTO path_family_members(family_id, path_hash) VALUES (?, ?)",
            [(family["family_id"], path_hash) for family in sorted(path_families.values(), key=lambda item: item["pattern_path"]) for path_hash in sorted(family["member_path_hashes"])],
        )
        connection.commit()


def _path_metadata(logs: list[Any]) -> tuple[dict[str, dict[str, str]], dict[str, str | None]]:
    by_path: dict[str, dict[str, str]] = {}
    timestamps: list[str] = []
    for raw_log in logs:
        if not isinstance(raw_log, Mapping):
            continue
        observed_at = _normalize_datetime_text(raw_log.get("t_detect_time"))
        if observed_at:
            timestamps.append(observed_at)
        rule_code = _cell_text(raw_log.get("str_rule_id"))
        candidates = (
            ("str_process_full", "str_process_short", "str_md5"),
            ("str_parent_path_full", None, "str_parent_md5"),
            ("str_suspicious_process_ancestor_full", "str_suspicious_process_ancestor_short", "str_suspicious_process_ancestor_md5"),
        )
        for path_key, name_key, md5_key in candidates:
            raw_path = raw_log.get(path_key)
            if not isinstance(raw_path, str) or not raw_path.strip():
                continue
            try:
                normalized_path = normalize_windows_path(raw_path)
            except ValueError:
                continue
            item = by_path.setdefault(normalized_path, {})
            _set_if_text(item, "observed_at", observed_at)
            _set_if_text(item, "process_name", raw_log.get(name_key) if name_key else ntpath.basename(normalized_path))
            md5 = _cell_text(raw_log.get(md5_key))
            if md5 and _MD5_PATTERN.fullmatch(md5.lower()):
                item["md5"] = md5.lower()
            _set_if_text(item, "rule_code", rule_code)
    return by_path, {
        "first_seen_at": min(timestamps) if timestamps else None,
        "last_seen_at": max(timestamps) if timestamps else None,
    }


def _is_temporary_path(path: str) -> bool:
    return bool(re.search(r"(^[a-z]:\\temp\\|\\windows\\temp\\|\\appdata\\local\\temp\\)", path))


def _is_user_writable_path(path: str) -> bool:
    return bool(re.search(r"^[a-z]:\\users\\|\\(desktop|downloads|documents|appdata)\\", path))


def _drive(path: str) -> str | None:
    drive, _ = ntpath.splitdrive(path)
    return drive.upper() if drive else None


def _path_warnings(control_zone: PingAnSoftwarePathControlZone) -> list[str]:
    if control_zone in {
        PingAnSoftwarePathControlZone.LESS_MANAGED,
        PingAnSoftwarePathControlZone.USER_WRITABLE,
        PingAnSoftwarePathControlZone.TEMPORARY,
    }:
        return ["Path is in a PingAn higher-attention location; historical occurrence does not remove location risk."]
    if control_zone in {
        PingAnSoftwarePathControlZone.MANAGED_PROGRAM,
        PingAnSoftwarePathControlZone.MANAGED_SYSTEM,
    }:
        return ["Managed location is only a lowering context signal; executable behavior and invocation still require analysis."]
    return []


def _freshness(
    last_seen: datetime | None,
    *,
    as_of: datetime | None,
    freshness_days: int,
) -> PingAnSoftwarePathFreshness:
    if last_seen is None:
        return PingAnSoftwarePathFreshness.UNKNOWN
    reference = as_of or datetime.now(UTC)
    if reference.tzinfo is None:
        reference = reference.replace(tzinfo=UTC)
    return PingAnSoftwarePathFreshness.CURRENT if (reference - last_seen).days <= freshness_days else PingAnSoftwarePathFreshness.STALE


def _parse_datetime(value: Any) -> datetime | None:
    if not isinstance(value, str) or not value:
        return None
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        return None
    return parsed if parsed.tzinfo is not None else parsed.replace(tzinfo=UTC)


def _normalize_datetime_text(value: Any) -> str | None:
    text = _cell_text(value)
    if not text:
        return None
    for parser in (datetime.fromisoformat, lambda item: datetime.strptime(item, "%Y-%m-%d %H:%M:%S")):
        try:
            parsed = parser(text)
        except ValueError:
            continue
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=UTC)
        return parsed.isoformat()
    return None


def _update_seen_range(entry: dict[str, Any], observed_at: str | None) -> None:
    if not observed_at:
        return
    if entry["first_seen_at"] is None or observed_at < entry["first_seen_at"]:
        entry["first_seen_at"] = observed_at
    if entry["last_seen_at"] is None or observed_at > entry["last_seen_at"]:
        entry["last_seen_at"] = observed_at


def _minimum_text(values: Iterable[Any]) -> str | None:
    normalized = [text for value in values if (text := _cell_text(value))]
    return min(normalized) if normalized else None


def _maximum_text(values: Iterable[Any]) -> str | None:
    normalized = [text for value in values if (text := _cell_text(value))]
    return max(normalized) if normalized else None


def _non_empty_text_set(values: Iterable[Any]) -> set[str]:
    return {text for value in values if (text := _cell_text(value))}


def _parse_json_object(value: Any) -> dict[str, Any] | None:
    if isinstance(value, Mapping):
        return dict(value)
    if not isinstance(value, str):
        return None
    try:
        parsed = json.loads(value)
    except (TypeError, ValueError):
        return None
    return dict(parsed) if isinstance(parsed, Mapping) else None


def _parse_json_list(value: Any) -> list[Any] | None:
    if isinstance(value, list):
        return value
    if not isinstance(value, str):
        return None
    try:
        parsed = json.loads(value)
    except (TypeError, ValueError):
        return None
    return parsed if isinstance(parsed, list) else None


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text or None


def _set_if_text(target: dict[str, str], key: str, value: Any) -> None:
    text = _cell_text(value)
    if text and key not in target:
        target[key] = text


def _add_if_text(target: set[str], value: Any) -> None:
    text = _cell_text(value)
    if text:
        target.add(text)


def _json_set(values: Iterable[Any], *, limit: int | None = None) -> str:
    items = sorted({_cell_text(value) for value in values if _cell_text(value)})
    if limit is not None:
        items = items[:limit]
    return json.dumps(items, ensure_ascii=False, separators=(",", ":"))


def _json_string_list(value: Any, *, limit: int | None = None) -> list[str]:
    try:
        loaded = json.loads(str(value))
    except (TypeError, ValueError):
        return []
    if not isinstance(loaded, list):
        return []
    result = [str(item) for item in loaded if isinstance(item, str) and item]
    return result[:limit] if limit is not None else result


def _required_metadata(metadata: Mapping[str, str], key: str) -> str:
    value = metadata.get(key)
    if not value:
        raise PingAnSoftwarePathCatalogError(f"catalog metadata is missing {key}")
    return value


def _required_payload_string(payload: Mapping[str, Any], key: str) -> str:
    value = payload.get(key)
    if not isinstance(value, str) or not value.strip():
        raise SocActionAdapterRegistryError(f"endpoint.software_path.lookup requires non-empty payload {key}")
    return value.strip()


def _normalize_optional_md5(value: str | None) -> str | None:
    if value is None or not value.strip():
        return None
    normalized = value.strip().lower()
    if not _MD5_PATTERN.fullmatch(normalized):
        raise ValueError("md5 must be a 32-character hexadecimal value")
    return normalized


def _sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _count_values(values: Iterable[Any]) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for value in values:
        counts[str(value)] += 1
    return dict(sorted(counts.items()))


__all__ = [
    "DEFAULT_CATALOG_ENV",
    "PINGAN_SOFTWARE_PATH_LOOKUP_ACTION",
    "PingAnSoftwarePathAttention",
    "PingAnSoftwarePathCatalog",
    "PingAnSoftwarePathCatalogBuildReport",
    "PingAnSoftwarePathCatalogError",
    "PingAnSoftwarePathContextResult",
    "PingAnSoftwarePathControlZone",
    "PingAnSoftwarePathFreshness",
    "PingAnSoftwarePathFamilyContext",
    "PingAnSoftwarePathHistoricalContext",
    "PingAnSoftwarePathLookupActionAdapter",
    "PingAnSoftwarePathMatchType",
    "classify_pingan_path",
    "compile_pingan_software_path_catalog",
    "is_executable_like_path",
    "normalize_windows_path",
    "pingan_software_path_adapter_descriptor",
]
